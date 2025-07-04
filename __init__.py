import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from .core import Beluga, create_question_df
from .config import BelugaConfig
from .errors import BelugaValidationError
from .utils.html import *
from .utils.jsLogic import *
from .utils import *
from .belugaQuestion import *
from .previewQuestion import *

def create_question_df() :
    """설문 문항을 저장할 DataFrame을 생성합니다."""
    import pandas as pd
    return pd.DataFrame(columns=[
            'QID',
            '문항번호',
            '문항유형',
            '최소',
            '최대',
            '보기 순서 로테이션',
            '보기 파이핑 적용',
            '질문',
            '보기',
            '조건',
            '실패시 이동',
            '응답후 로직',
            '단위',
        ])


class BelugaConfig:
    """Beluga 클래스의 설정을 관리하는 클래스입니다."""
    def __init__(self, etc_text: str = '기타(직접 입력)', default_rotation: bool = False, dropdown_placeholder: str = '하나 선택...', total_text: str = '합계', display: bool = True, change: bool = True):
        self.etc_text = etc_text
        self.dropdown_placeholder = dropdown_placeholder
        self.total_text = total_text
        self.default_rotation = default_rotation
        self.display = display
        self.change = change

class BelugaValidationError(Exception):
    """Beluga 클래스 검증 에러"""
    pass

class Beluga:
    """
    설문 문항을 생성하고 관리하는 클래스입니다.
    다양한 유형의 문항을 DataFrame 형태로 관리할 수 있습니다.
    """
    ALLOWED_SCALE_SCORES = [3, 4, 5, 6, 7, 9, 10, 11]
    DEFAULT_ETC_TEXT = '기타(직접 입력)'

    def __init__(self, config: Optional[BelugaConfig] = None) -> None:
        """
        Beluga 인스턴스를 초기화합니다.

        Args:
            config (Optional[BelugaConfig]): 설정 객체. None인 경우 기본 설정 사용
        """
        self.config = config or BelugaConfig()
        self.df = create_question_df()
        self.etc_text = self.config.etc_text
        self.dropdown_placeholder = self.config.dropdown_placeholder
        self.total_text = self.config.total_text
        self._qid_cache = set()

    def set_etc_text(self, new_etc_text: str) -> None:
        """
        기타 선택지 텍스트를 변경합니다.

        Args:
            new_etc_text (str): 새로운 기타 선택지 텍스트
        """
        self.etc_text = new_etc_text
        self.config.etc_text = new_etc_text

    def _validate_question(self, qid: Optional[str], title: str, min_val: Optional[int], max_val: Optional[int], change: bool = True) -> None:
        """
        문항 정보의 유효성을 검증합니다.

        Args:
            qid (Optional[str]): 질문 ID
            title (str): 질문 제목
            min_val (Optional[int]): 최소 선택 개수
            max_val (Optional[int]): 최대 선택 개수
            change (bool): 기존 QID 변경 여부

        Raises:
            BelugaValidationError: 검증 실패시 발생
        """
        if not title:
            raise BelugaValidationError("제목은 필수입니다.")
        if min_val is not None and max_val is not None:
            if min_val > max_val:
                raise BelugaValidationError("최소 선택 개수는 최대 선택 개수보다 작아야 합니다.")
            if min_val < 0:
                raise BelugaValidationError("최소 선택 개수는 0 이상이어야 합니다.")
        if qid is not None and qid in self._qid_cache and not change:
            raise BelugaValidationError("QID가 중복됩니다. 내용을 변경하는 것이라면 change=True로 설정하세요.")

    def _format_options(self, options: Union[dict, list[str]], na: Optional[str] = None, etc_text: str = None) -> str:
        """
        선택지를 포맷팅합니다.

        Args:
            options (Union[dict, str]): 선택지 목록 또는 문자열
            na (Optional[str]): 무응답 옵션
            etc_text (Optional[str]): 기타 선택지 텍스트

        Returns:
            str: 포맷팅된 선택지 문자열
        """
        if not options:
            return ""
        if isinstance(options, dict):
            opts = {0: na, **options} if na is not None else options
            formatted = []
            for idx, opt in opts.items():
                if opt == (self.etc_text if etc_text is None else etc_text):
                    formatted.append(f"{idx}E) {opt}")
                else:
                    formatted.append(f"{idx}) {opt}")
            return '\n'.join(formatted)
        elif isinstance(options, list) :
            opts = [na, *options] if na is not None else options
            formatted = []
            for idx, opt in enumerate(opts, 0 if na is not None else 1):
                if opt == (self.etc_text if etc_text is None else etc_text):
                    formatted.append(f"{idx}E) {opt}")
                else:
                    formatted.append(f"{idx}) {opt}")
            return '\n'.join(formatted)
        else :
            return str(options)

    def show_df(self, qid: str = None) :
        """
        DataFrame을 조회합니다.

        Args:
            qid (str, optional): 특정 QID의 문항만 조회. None인 경우 전체 조회

        Returns:
            pd.DataFrame: 조회된 DataFrame
        """
        if qid is not None :
            return self.df[self.df['QID'] == qid]
        return self.df

    def show_options(self, qid: str) -> None :
        """
        특정 QID의 선택지를 출력합니다.

        Args:
            qid (str): 질문 ID
        """
        for i in self.df[self.df['QID'] == qid]['보기'].tolist() :
            print(i)


    def sa(
        self,
        title: str,
        qid: Optional[str] = None,
        options: Union[dict, list[str]] = {},
        scale: bool = False,
        etc: bool = False,
        etc_text: str = None,
        na: Optional[str] = None,
        cond: Optional[Union[str, list[str]]] = None,
        piping: Union[str, int, BelugaQuestion] = None,
        selected_piping: bool = True,
        rotation: Optional[bool] = None,
        fail: str = '',
        post_logic: str = '',
        change: Optional[bool] = None,
        inplace: bool = True,
    ) -> 'Beluga' :
        """
        단일 선택 객관식 문항을 추가합니다.

        Args:
            title (str): 질문 제목
            qid (Optional[str]): 질문 ID. None인 경우 자동 생성
            options (Union[List[str], str]): 선택지 목록 또는 문자열
            scale (bool): 척도형 문항 여부
            etc (bool): 기타 선택지 포함 여부
            etc_text (str): 기타 선택지 텍스트. None인 경우 기본값 사용
            na (Optional[str]): 무응답 옵션
            cond (Optional[Union[str, list[str]]]): 조건문
            piping (Union[str, int, BelugaQuestion]): 파이핑 설정
            rotation (Optional[bool]): 보기 순서 로테이션 여부. None인 경우 기본값 사용
            fail (str): 조건 실패시 이동할 위치
            post_logic (str): 응답 후 로직
            change (bool): 기존 QID 변경 여부
            inplace (bool): 현재 인스턴스에 추가할지 여부

        Returns:
            Beluga: qid가 지정된 경우 해당 질문의 DataFrame, 아니면 마지막 추가된 질문의 DataFrame
        """
        rotation_val = self.config.default_rotation if rotation is None else rotation
        change_val = self.config.change if change is None else change
        if scale :
            rotation_val = False
            etc = False
            na = None

        self._validate_question(qid, title, None, None, change_val)
        return self.append_question(
            qid=qid,
            qtype='객관식 단일' if not scale else '객관식 척도',
            title=title,
            options=options,
            etc=etc,
            etc_text=etc_text,
            na=na,
            cond=cond,
            piping=piping,
            selected_piping=selected_piping,
            rotation=rotation_val,
            fail=fail,
            post_logic=post_logic,
            change=change_val,
            inplace=inplace
        )

    def ma(
        self,
        title: str,
        qid: Optional[str] = None,
        options: Union[dict, list[str]] = {},
        etc: bool = False,
        etc_text: str = None,
        na: Optional[str] = None,
        cond: Optional[Union[str, list[str]]] = None,
        min: Optional[int] = 1,
        max: Optional[int] = None,
        piping: Union[str, int, BelugaQuestion] = None,
        selected_piping: bool = True,
        rotation: Optional[bool] = None,
        fail: str = '',
        post_logic: str = '',
        change: Optional[bool] = None,
        inplace: bool = True,
    ) -> 'Beluga':
        """
        다중 선택 객관식 문항을 추가합니다.

        Args:
            title (str): 질문 제목
            qid (Optional[str]): 질문 ID. None인 경우 자동 생성
            options (Union[List[str], str]): 선택지 목록 또는 문자열
            etc (bool): 기타 선택지 포함 여부
            etc_text (str): 기타 선택지 텍스트. None인 경우 기본값 사용
            na (Optional[str]): 무응답 옵션
            cond (Optional[Union[str, list[str]]]): 조건문
            min (Optional[int]): 최소 선택 개수
            max (Optional[int]): 최대 선택 개수
            piping (Union[str, int, BelugaQuestion]): 파이핑 설정
            rotation (Optional[bool]): 보기 순서 로테이션 여부. None인 경우 기본값 사용
            fail (str): 조건 실패시 이동할 위치
            post_logic (str): 응답 후 로직
            change (bool): 기존 QID 변경 여부
            inplace (bool): 현재 인스턴스에 추가할지 여부

        Returns:
            Beluga: qid가 지정된 경우 해당 질문의 DataFrame, 아니면 마지막 추가된 질문의 DataFrame
        """
        rotation_val = self.config.default_rotation if rotation is None else rotation
        change_val = self.config.change if change is None else change
        self._validate_question(qid, title, min, max, change_val)

        if max is None :
            max = len(options)
            if etc :
                max += 1

        return self.append_question(
            qid=qid,
            qtype='객관식 중복',
            title=title,
            options=options,
            etc=etc,
            etc_text=etc_text,
            na=na,
            cond=cond,
            min=min,
            max=max,
            piping=piping,
            selected_piping=selected_piping,
            rotation=rotation_val,
            fail=fail,
            post_logic=post_logic,
            change=change_val,
            inplace=inplace
        )


    def rank(
        self,
        title: str,
        qid: Optional[str] = None,
        options: Union[dict, list[str]] = {},
        etc: bool = False,
        etc_text: str = None,
        na: Optional[str] = None,
        cond: Optional[Union[str, list[str]]] = None,
        min: Optional[int] = 1,
        max: Optional[int] = None,
        piping: Union[str, int, BelugaQuestion] = None,
        selected_piping: bool = True,
        rotation: Optional[bool] = None,
        fail: str = '',
        post_logic: str = '',
        change: Optional[bool] = None,
        inplace: bool = True,
    ) -> 'Beluga':
        """
        순위 선택 객관식 문항을 추가합니다.

        Args:
            title (str): 질문 제목
            qid (Optional[str]): 질문 ID. None인 경우 자동 생성
            options (Union[List[str], str]): 선택지 목록 또는 문자열
            etc (bool): 기타 선택지 포함 여부
            etc_text (str): 기타 선택지 텍스트. None인 경우 기본값 사용
            na (Optional[str]): 무응답 옵션
            cond (Optional[Union[str, list[str]]]): 조건문
            min (Optional[int]): 최소 선택 개수
            max (Optional[int]): 최대 선택 개수
            piping (Union[str, int, BelugaQuestion]): 파이핑 설정
            rotation (Optional[bool]): 보기 순서 로테이션 여부. None인 경우 기본값 사용
            fail (str): 조건 실패시 이동할 위치
            post_logic (str): 응답 후 로직
            change (bool): 기존 QID 변경 여부
            inplace (bool): 현재 인스턴스에 추가할지 여부

        Returns:
            Beluga: qid가 지정된 경우 해당 질문의 DataFrame, 아니면 마지막 추가된 질문의 DataFrame
        """
        rotation_val = self.config.default_rotation if rotation is None else rotation
        change_val = self.config.change if change is None else change
        self._validate_question(qid, title, min, max, change_val)

        if max is None :
            max = len(options)
            if etc :
                max += 1

        return self.append_question(
            qid=qid,
            qtype='객관식 순위',
            title=title,
            options=options,
            etc=etc,
            etc_text=etc_text,
            na=na,
            cond=cond,
            min=min,
            max=max,
            piping=piping,
            selected_piping=selected_piping,
            rotation=rotation_val,
            fail=fail,
            post_logic=post_logic,
            change=change_val,
            inplace=inplace
        )


    def text(
        self,
        title: str,
        qid: Optional[str] = None,
        cond: Optional[Union[str, list[str]]] = None,
        fail: str = '',
        post_logic: str = '',
        change: Optional[bool] = None,
        multi: Optional[Union[int, list[str]]] = None,
        multi_atleast: bool = False,
        multi_post: str = None,
        multi_width: str = '200px',
        inplace: bool = True,
    ) -> 'Beluga':
        """
        주관식 문자 입력 문항을 추가합니다.

        Args:
            title (str): 질문 제목
            qid (Optional[str]): 질문 ID. None인 경우 자동 생성
            cond (Optional[Union[str, list[str]]]): 조건문
            fail (str): 조건 실패시 이동할 위치
            post_logic (str): 응답 후 로직
            change (bool): 기존 QID 변경 여부
            multi (Optional[Union[int, list[str]]]): 다중 입력 설정 (개수 또는 라벨 목록)
            multi_atleast (bool): 다중 입력시 최소 하나 이상 입력 필수 여부
            multi_post (str): 다중 입력 후 텍스트
            multi_width (str): 입력 필드 너비
            inplace (bool): 현재 인스턴스에 추가할지 여부

        Returns:
            Beluga: 추가된 문항의 DataFrame
        """
        change_val = self.config.change if change is None else change
        if multi is not None :
            if isinstance(multi, (int, list, dict)) :
                multi_html = set_multi_input(n=multi, type='text', post_text=multi_post, width=multi_width)
            else :
                raise BelugaValidationError("multi는 int, list[str], dict 형식이어야 합니다.")
            title = f'{title}\n{multi_html}'
            js = None
            if multi_atleast :
                js = multi_text_atleast_js
            else :
                js = multi_text_all_js
            if cond is not None :
                if isinstance(cond, list) :
                    cond.append(js)
                else :
                    cond += f' && {js}'
            else :
                cond = js

        return self.append_question(
            qid=qid,
            qtype='주관식 문자',
            title=title,
            cond=cond,
            fail=fail,
            post_logic=post_logic,
            change=change_val,
            inplace=inplace
        )


    def address(
        self,
        title: str,
        qid: Optional[str] = None,
        cond: Optional[Union[str, list[str]]] = None,
        fail: str = '',
        post_logic: str = '',
        change: Optional[bool] = None,
        inplace: bool = True,
    ) -> 'Beluga':
        """
        주관식 주소 입력 문항을 추가합니다.

        Args:
            title (str): 질문 제목
            qid (Optional[str]): 질문 ID. None인 경우 자동 생성
            cond (Optional[Union[str, list[str]]]): 조건문
            fail (str): 조건 실패시 이동할 위치
            post_logic (str): 응답 후 로직
            change (bool): 기존 QID 변경 여부
            inplace (bool): 현재 인스턴스에 추가할지 여부

        Returns:
            Beluga: 추가된 문항의 DataFrame
        """
        change_val = self.config.change if change is None else change
        return self.append_question(
            qid=qid,
            qtype='주관식 주소',
            title=title,
            cond=cond,
            fail=fail,
            post_logic=post_logic,
            change=change_val,
            inplace=inplace
        )

    def phone(
        self,
        title: str,
        qid: Optional[str] = None,
        cond: Optional[Union[str, list[str]]] = None,
        fail: str = '',
        post_logic: str = '',
        change: Optional[bool] = None,
        inplace: bool = True,
    ) -> 'Beluga':
        """
        주관식 전화번호 입력 문항을 추가합니다.

        Args:
            title (str): 질문 제목
            qid (Optional[str]): 질문 ID. None인 경우 자동 생성
            cond (Optional[Union[str, list[str]]]): 조건문
            fail (str): 조건 실패시 이동할 위치
            post_logic (str): 응답 후 로직
            change (bool): 기존 QID 변경 여부
            inplace (bool): 현재 인스턴스에 추가할지 여부

        Returns:
            Beluga: 추가된 문항의 DataFrame
        """
        change_val = self.config.change if change is None else change
        return self.append_question(
            qid=qid,
            qtype='주관식 전화번호',
            title=title,
            cond=cond,
            fail=fail,
            post_logic=post_logic,
            change=change_val,
            inplace=inplace
        )


    def date(
        self,
        title: str,
        qid: Optional[str] = None,
        cond: Optional[Union[str, list[str]]] = None,
        fail: str = '',
        post_logic: str = '',
        change: Optional[bool] = None,
        inplace: bool = True,
    ) -> 'Beluga':
        """
        주관식 날짜 입력 문항을 추가합니다.

        Args:
            title (str): 질문 제목
            qid (Optional[str]): 질문 ID. None인 경우 자동 생성
            cond (Optional[Union[str, list[str]]]): 조건문
            fail (str): 조건 실패시 이동할 위치
            post_logic (str): 응답 후 로직
            change (bool): 기존 QID 변경 여부
            inplace (bool): 현재 인스턴스에 추가할지 여부

        Returns:
            Beluga: 추가된 문항의 DataFrame
        """
        change_val = self.config.change if change is None else change
        return self.append_question(
            qid=qid,
            qtype='주관식 날짜',
            title=title,
            cond=cond,
            fail=fail,
            post_logic=post_logic,
            change=change_val,
            inplace=inplace
        )


    def number(
        self,
        title: str,
        qid: Optional[str] = None,
        cond: Optional[Union[str, list[str]]] = None,
        min: Optional[int] = None,
        max: Optional[int] = None,
        total: Optional[int] = None,
        fail: str = '',
        post_logic: str = '',
        post_text: str = '',
        change: Optional[bool] = None,
        multi: Optional[Union[int, list[str]]] = None,
        multi_post: str = None,
        multi_width: str = '70px',
        inplace: bool = True,
    ) -> 'Beluga':
        """
        주관식 숫자 입력 문항을 추가합니다.

        Args:
            title (str): 질문 제목
            qid (Optional[str]): 질문 ID. None인 경우 자동 생성
            cond (Optional[Union[str, list[str]]]): 조건문
            min (Optional[int]): 최소값 (필수)
            max (Optional[int]): 최대값 (필수)
            total (Optional[int]): 합계 제한값 (multi 사용시)
            post_text (str): 입력 후 텍스트
            fail (str): 조건 실패시 이동할 위치
            post_logic (str): 응답 후 로직
            change (bool): 기존 QID 변경 여부
            multi (Optional[Union[int, list[str]]]): 다중 입력 설정 (개수 또는 라벨 목록)
            multi_post (str): 다중 입력 후 텍스트
            multi_width (str): 입력 필드 너비
            inplace (bool): 현재 인스턴스에 추가할지 여부

        Returns:
            Beluga: 추가된 문항의 DataFrame
        """
        change_val = self.config.change if change is None else change
        if any(i is None for i in [min, max]) :
            raise BelugaValidationError("min, max 반드시 입력해야 합니다.")

        if total is not None and multi is None :
            raise BelugaValidationError("multi 형태로 변경 필요")

        qtype = '주관식 숫자'

        if multi is not None :
            qtype = '주관식 문자'
            if isinstance(multi, (int, list, dict)) :
                multi_html = set_multi_input(n=multi, type='number', post_text=multi_post, width=multi_width, total=total is not None, total_label=self.total_text)
            else :
                raise BelugaValidationError("multi는 int, list[str], dict 형식이어야 합니다.")

            title = f'{title}\n{multi_html}'

            js = multi_num_js.format(min=min, max=max, total=total if total is not None else 'null')

            if cond is not None :
                if isinstance(cond, list) :
                    cond.append(js)
                else :
                    cond += f' && {js}'
            else :
                cond = js

        return self.append_question(
            qid=qid,
            qtype=qtype,
            title=title,
            min=min,
            max=max,
            cond=cond,
            post_text=post_text,
            fail=fail,
            post_logic=post_logic,
            change=change_val,
            inplace=inplace
        )

    def scale(
        self,
        title: str,
        qid: Optional[str] = None,
        left: str = '',
        center: str = '',
        right: str = '',
        score: int = 5,
        cond: Optional[Union[str, list[str]]] = None,
        fail: str = '',
        post_logic: str = '',
        change: Optional[bool] = None,
        inplace: bool = True,
    ) -> 'Beluga':
        """
        평가형(척도형) 문항을 추가합니다.
        허용 점수는 ALLOWED_SCALE_SCORES에 정의되어 있습니다.

        Args:
            title (str): 질문 제목
            qid (Optional[str]): 질문 ID. None인 경우 자동 생성
            left (str): 왼쪽 라벨 (최저점)
            center (str): 중간 라벨
            right (str): 오른쪽 라벨 (최고점)
            score (int): 척도 점수 (3, 4, 5, 6, 7, 9, 10, 11 중 선택)
            cond (Optional[Union[str, list[str]]]): 조건문
            fail (str): 조건 실패시 이동할 위치
            post_logic (str): 응답 후 로직
            change (bool): 기존 QID 변경 여부
            inplace (bool): 현재 인스턴스에 추가할지 여부

        Returns:
            Beluga: 추가된 문항의 DataFrame
        """
        if score not in self.ALLOWED_SCALE_SCORES:
            raise BelugaValidationError(f"score는 {self.ALLOWED_SCALE_SCORES} 중 하나여야 합니다.")
        qtype = f'평가형 {score}점'
        min_score = 0 if score == 11 else 1
        max_score = 10 if score == 11 else score
        center_score = score // 2 + 1
        options = [
            f'{min_score}) {left}',
            f'{center_score}) {center}',
            f'{score}) {right}'
        ]
        options = '\n'.join(options)
        change_val = self.config.change if change is None else change
        self._validate_question(qid, title, None, None, change_val)
        return self.append_question(
            qid=qid,
            qtype=qtype,
            title=title,
            options=options,
            cond=cond,
            fail=fail,
            post_logic=post_logic,
            change=change_val,
            inplace=inplace
        )


    def image(
        self,
        title: str,
        qid: Optional[str] = None,
        cond: Optional[Union[str, list[str]]] = None,
        fail: str = '',
        post_logic: str = '',
        change: Optional[bool] = None,
        inplace: bool = True,
    ) -> 'Beluga':
        """
        이미지 업로드 문항을 추가합니다.

        Args:
            title (str): 질문 제목
            qid (Optional[str]): 질문 ID. None인 경우 자동 생성
            cond (Optional[Union[str, list[str]]]): 조건문
            fail (str): 조건 실패시 이동할 위치
            post_logic (str): 응답 후 로직
            change (bool): 기존 QID 변경 여부
            inplace (bool): 현재 인스턴스에 추가할지 여부

        Returns:
            Beluga: 추가된 문항의 DataFrame
        """
        change_val = self.config.change if change is None else change
        return self.append_question(
            qid=qid,
            qtype='이미지',
            title=title,
            cond=cond,
            fail=fail,
            post_logic=post_logic,
            change=change_val,
            inplace=inplace
        )

    def dropdown(
        self,
        title: str,
        qid: Optional[str] = None,
        cond: Optional[Union[str, list[str]]] = None,
        options: Union[dict, list[str]] = {},
        rows: Union[dict, list[str]] = {},
        row_cond: Optional[Union[str, int]] = None,
        duplicate: bool = False,
        fail: str = '',
        post_logic: str = '',
        change: Optional[bool] = None,
        inplace: bool = True,
    ) -> 'Beluga':
        if not isinstance(options, (dict, list)) or len(options) == 0:
            raise BelugaValidationError("options는 빈 리스트가 아니어야 합니다.")

        if not isinstance(rows, (dict, list)) or len(rows) == 0:
            raise BelugaValidationError("rows는 빈 리스트가 아니어야 합니다.")

        change_val = self.config.change if change is None else change
        selects = set_dropdown(options=options, rows=rows, placeholder=self.dropdown_placeholder)
        title = f'{title}\n{selects}'

        if row_cond is not None :
            if isinstance(row_cond, str) :
                base_qnum = self.df[self.df.QID == row_cond]['문항번호'].iloc[0]
                if base_qnum is None :
                    raise BelugaValidationError(f"row_cond에 해당하는 문항이 없습니다. {row_cond}")
            elif isinstance(row_cond, int) :
                base_qnum = row_cond
            else :
                raise BelugaValidationError("row_cond는 str 또는 int 형식이어야 합니다.")

            row_cond_js = dropdown_row_cond.format(base=base_qnum)

            if cond is not None :
                if isinstance(cond, list) :
                    cond.append(row_cond_js)
                else :
                    cond += f' && {row_cond_js}'
            else :
                cond = row_cond_js

        if cond is not None :
            js = dropdown_js.format(duplicate= 'true' if duplicate else 'false')
            if isinstance(cond, list) :
                cond.append(js)
            else :
                cond += f' && {js}'
        else :
            cond = js


        return self.append_question(
            qid=qid,
            qtype='주관식 문자',
            title=title,
            cond=cond,
            fail=fail,
            post_logic=post_logic,
            change=change_val,
            inplace=inplace
        )

    def append_question(
        self,
        qtype: str,
        title: str,
        options: Union[dict, list[str], str] = {},
        etc: bool = False,
        etc_text: str = None,
        na: Optional[str] = None,
        cond: Optional[Union[str, list[str]]] = None,
        min: Optional[int] = None,
        max: Optional[int] = None,
        piping: Union[str, int, BelugaQuestion] = None,
        selected_piping: bool = True,
        rotation: bool = False,
        fail: str = '',
        post_logic: str = '',
        post_text: str = '',
        qid: Optional[str] = None,
        change: bool = True,
        inplace: bool = True,
    ) -> pd.DataFrame:
        """
        문항 정보를 DataFrame에 추가하거나, 새 DataFrame을 반환합니다.

        Args:
            qtype (str): 문항 유형
            title (str): 질문 제목
            options (Union[dict, str]): 선택지 목록 또는 문자열
            etc (bool): 기타 선택지 포함 여부
            etc_text (str): 기타 선택지 텍스트. None인 경우 기본값 사용
            na (Optional[str]): 무응답 옵션
            cond (Optional[Union[str, list[str]]]): 조건문
            min (Optional[int]): 최소값/최소 선택 개수
            max (Optional[int]): 최대값/최대 선택 개수
            piping (Union[str, int, BelugaQuestion]): 파이핑 설정
            rotation (bool): 보기 순서 로테이션 여부
            fail (str): 조건 실패시 이동할 위치
            post_logic (str): 응답 후 로직
            post_text (str): 응답 후 텍스트
            qid (Optional[str]): 질문 ID
            change (bool): 기존 QID 변경 여부
            inplace (bool): 현재 인스턴스에 추가할지 여부

        Returns:
            pd.DataFrame: 추가된 문항의 DataFrame
        """

        if isinstance(options, (list, dict)) :
            check_group = group_rot(options)
            if check_group is not None :
                rotation = True
                if cond is not None :
                    if isinstance(cond, list) :
                        cond.append(check_group)
                    else :
                        cond = f'{cond} && {check_group}'
                else :
                    cond = check_group


                if isinstance(options, list) :
                    # Flatten nested lists
                    flattened = []
                    for item in options:
                        if isinstance(item, list):
                            flattened.extend(item)
                        else:
                            flattened.append(item)
                    options = flattened
                    options = [i if isinstance(i, str) else i.keys() for i in options]

                elif isinstance(options, dict) :
                    # Flatten nested dicts
                    flattened = {}
                    for key, value in options.items():
                        if isinstance(value, dict):
                            flattened.update(value)
                        else:
                            flattened[key] = value
                    options = flattened


        if etc :
            if isinstance(options, dict) and options:
                keys = list(options.keys())
                max_key = sorted([int(k) for k in keys])[-1]
                options[max_key + 1] = self.etc_text if etc_text is None else etc_text

            elif isinstance(options, list) :
                options.append(self.etc_text if etc_text is None else etc_text)

        # 조건 처리
        if cond is not None:
            if isinstance(cond, list):
                cond = ' && '.join(cond)
        else:
            cond = ''

        if piping is not None :
            if isinstance(piping, BelugaQuestion) :
                piping = f'Q{piping.qnum}'
            elif isinstance(piping, int) :
                piping = f'Q{piping}'
            elif isinstance(piping, str) :
                qnum = self.df[self.df.QID == piping]['문항번호'].iloc[0]
                if qnum is None :
                    raise BelugaValidationError(f"piping에 해당하는 문항이 없습니다. {piping}")
                piping = f'Q{qnum}'

            if not selected_piping :
                piping = f'!{piping}'
        else :
            piping = ''

        if isinstance(options, (list, dict)) :
            options = self._format_options(options, na, etc_text)

        if cond is not None :
            cond = cond.strip()


        for text, var_name in [(title, 'title'), (cond, 'cond')]:
            for match in extract_qids(text):
                qnum_ref = self.df[self.df.QID == match]['문항번호'].iloc[0]
                if qnum_ref is not None:
                    if var_name == 'title':
                        title = title.replace(f'#[{match}]', str(qnum_ref))
                    else:
                        cond = cond.replace(f'#[{match}]', str(qnum_ref))


        title = title.strip()
        if cond:
            cond = cond.strip()

        if inplace:
            self._append_inplace(
                qtype, title, options, na, cond, min, max, piping, rotation, fail, post_logic, post_text, qid, change, etc_text
            )
        else:
            self._create_dummy_df(
                qtype, title, options, na, cond, min, max, piping, rotation, fail, post_logic, post_text, qid, etc_text
            )

        # Attributes
        if qid is not None :
            if not change and hasattr(self, qid) :
                raise BelugaValidationError(f"QID가 중복됩니다. 내용을 변경하는 것이라면 change=True로 설정하세요.")
            else :
                qnum = self.df[self.df.QID == qid]['문항번호'].iloc[0]
                question = BelugaQuestion(
                    type=qtype,
                    qid=qid,
                    qnum=qnum,
                    title=title,
                    options=options,
                    na=na,
                    etc=etc,
                    etc_text=etc_text,
                    cond=cond,
                    min=min,
                    max=max,
                    piping=piping,
                    selected_piping=selected_piping,
                    rotation=rotation,
                    fail=fail,
                    post_logic=post_logic,
                    post_text=post_text,
                )
                setattr(self, qid, question)
        else :
            qnum = len(self.df.index)
            qid = f'Q{qnum}'
            question = BelugaQuestion(
                type=qtype,
                qid=qid,
                qnum=qnum,
                title=title,
                options=options,
                na=na,
                etc=etc,
                etc_text=etc_text,
                cond=cond,
                min=min,
                max=max,
                piping=piping,
                selected_piping=selected_piping,
                rotation=rotation,
                fail=fail,
                post_logic=post_logic,
                post_text=post_text,
            )

            setattr(self, qid, question)

        if self.config.display:
            preview_question(question)

        return getattr(self, qid)

    def _append_inplace(
        self, qtype, title, options, na, cond, min, max, piping, rotation, fail, post_logic, post_text, qid, change, etc_text
    ) -> pd.DataFrame:
        """
        현재 인스턴스의 DataFrame에 문항을 추가하거나 기존 문항을 업데이트합니다.

        Returns:
            pd.DataFrame: 추가되거나 업데이트된 문항의 DataFrame
        """
        # change=True이고 qid가 존재하는 경우 기존 row 업데이트
        if change and qid is not None and qid in self._qid_cache:
            df_idx = self.df[self.df['QID'] == qid].index[0]
            qnum = self.df.loc[df_idx, '문항번호']
        else:
            # 새로운 row 추가
            qnum = len(self.df.index) + 1
            df_idx = len(self.df.index)
            self.df.loc[df_idx, '문항번호'] = qnum

        self.df.loc[df_idx, '문항유형'] = qtype
        self.df.loc[df_idx, '질문'] = title
        self.df.loc[df_idx, '보기'] = options
        self.df.loc[df_idx, '조건'] = cond
        self.df.loc[df_idx, '보기 순서 로테이션'] = 'v' if rotation else ''
        self.df.loc[df_idx, '보기 파이핑 적용'] = piping
        self.df.loc[df_idx, '최소'] = min if min is not None else ''
        self.df.loc[df_idx, '최대'] = max if max is not None else ''
        self.df.loc[df_idx, '실패시 이동'] = fail
        self.df.loc[df_idx, '응답후 로직'] = post_logic
        self.df.loc[df_idx, '단위'] = post_text

        if qid is not None:
            if qid in self._qid_cache and not change:
                raise BelugaValidationError("QID가 중복됩니다. 내용을 변경하는 것이라면 change=True로 설정하세요.")
            self.df.loc[df_idx, 'QID'] = qid
            self._qid_cache.add(qid)

        if qid is not None:
            return self.show_df(qid)
        return self.df.iloc[-1:]

    def _create_dummy_df(
        self, qtype, title, options, na, cond, min, max, piping, rotation, fail, post_logic, post_text, qid, etc_text
    ) -> pd.DataFrame:
        """
        새로운 더미 DataFrame을 생성하여 문항 정보를 담아 반환합니다.

        Returns:
            pd.DataFrame: 문항 정보가 담긴 새 DataFrame
        """
        dummy_df = create_question_df()
        dummy_df.loc[0, '문항유형'] = qtype
        dummy_df.loc[0, '질문'] = title
        dummy_df.loc[0, '보기'] = options
        dummy_df.loc[0, '조건'] = cond
        dummy_df.loc[0, '보기 순서 로테이션'] = 'v' if rotation else ''
        dummy_df.loc[0, '보기 파이핑 적용'] = piping
        dummy_df.loc[0, '최소'] = min if min is not None else ''
        dummy_df.loc[0, '최대'] = max if max is not None else ''
        dummy_df.loc[0, '실패시 이동'] = fail
        dummy_df.loc[0, '응답후 로직'] = post_logic
        dummy_df.loc[0, '단위'] = post_text
        if qid is not None:
            dummy_df.loc[0, 'QID'] = qid
        return dummy_df



    def to_excel(self, path: str) -> None:
        """
        self.df를 지정된 엑셀 포맷으로 export합니다.
        """

        # 1. 데이터 준비 (QID 제외하지 않음)
        df = self.df[[
            'QID', '문항번호', '문항유형', '최소', '최대', '보기 순서 로테이션', '보기 파이핑 적용',
            '질문', '보기', '조건', '실패시 이동', '응답후 로직', '단위'
        ]]

        # 2. 새 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "Question"

        # 3. 헤더 정의 및 입력 (A열 제외)
        header_names = [
            '문항번호', '문항유형', '최소', '최대', '보기 순서 로테이션', '보기 파이핑 적용',
            '질문', '보기', '응답전로직\n(응답자 BASE)', '', '응답후 로직', '단위'
        ]
        header_sub = [
            '', '', '', '', '', '', '', '', '조건', '실패시\n이동', '', ''
        ]
        # B1~M1, B2~M2
        for col, (main, sub) in enumerate(zip(header_names, header_sub), start=2):
            ws.cell(row=1, column=col, value=main)
            ws.cell(row=2, column=col, value=sub)

        # 4. 셀 병합 (수직 병합: B1:B2, C1:C2, ..., I1:I2, L1:L2, M1:M2, J1:K1만 수평 병합)
        for col in list(range(2, 10)) + [12, 13]:  # B~I, L, M
            ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=11)  # J1:K1

        # 5. 헤더 스타일 적용
        header_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
        header_font = Font(size=10)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=2, max_col=13):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = align_center

        # 6. 데이터 입력 (A3~, B3~)
        for r, row in enumerate(df.values, start=3):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)

        # 7. 데이터 셀 스타일 (폰트, border, 위쪽 맞춤)
        align_top = Alignment(vertical='top', wrap_text=True)
        align_center_top = Alignment(horizontal='center', vertical='top', wrap_text=True)
        align_center_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # B, C, D, E, F, G = 2,3,4,5,6,7
        center_cols = [2, 3, 4, 5, 6, 7]
        for r in range(3, 3 + df.shape[0]):
            for c in range(1, 14):
                cell = ws.cell(row=r, column=c)
                if c == 1:  # A열 (QID)
                    cell.font = Font(size=10, bold=True, color='FF0000')
                    cell.alignment = align_center_top
                    # A열은 보더 적용하지 않음
                else:
                    cell.font = Font(size=10)
                    cell.border = border
                    if c in center_cols:
                        cell.alignment = align_center_top
                    else:
                        cell.alignment = align_top

        # 8. 열 너비 지정 및 자동조정
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 11
        ws.column_dimensions['D'].width = 5
        ws.column_dimensions['E'].width = 5
        ws.column_dimensions['F'].width = 9
        ws.column_dimensions['H'].width = 70
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 30
        ws.column_dimensions['L'].width = 7
        ws.column_dimensions['M'].width = 5
        # 나머지 열은 기존 자동조정 유지
        for col in range(2, 14):
            col_letter = get_column_letter(col)
            if col_letter not in ['B', 'C', 'D', 'E', 'F', 'H', 'I', 'J', 'L', 'M']:
                max_length = 0
                for row in ws.iter_rows(min_row=1, max_row=2+df.shape[0], min_col=col, max_col=col):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length+2, 50)

        # 9. 1행 높이 27
        ws.row_dimensions[1].height = 27

        wb.save(path)

        # 저장 후 엑셀 파일 자동 실행 (Windows)
        import os
        import sys
        if sys.platform.startswith('win'):
            os.startfile(path)


        cond_data = self.df[self.df['조건'].notna()]['조건'].tolist()
        js_code = []
        if cond_data :
            for key, js in pre_logic_dict.items() :
                append_flag = False
                if any(key in cond for cond in cond_data) :
                    append_flag = True
                    js_code.append(js)

                if not append_flag :
                    if key == 'validate' :
                        if any(any(c in cond for c in ['exec', 'cond', 'hangle']) for cond in cond_data) :
                            js_code.append(js)

                    if key == 'optionPosition' :
                        if any(any(c in cond for c in ['nextTo', 'beforeTo', 'topPosition']) for cond in cond_data) :
                            js_code.append(js)

        if js_code :
            with open('pre-logic.js', 'w', encoding='utf-8') as f :
                f.write('\n\n'.join(js_code))

